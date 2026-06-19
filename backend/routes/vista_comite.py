"""
Endpoints para Vista Comité con exportación Excel ejecutiva.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
from collections import defaultdict
import uuid
import io

# Openpyxl imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_router(db, get_current_user, CurrentUser):
    router = APIRouter()
    
    # ==========================================
    # MODELOS
    # ==========================================
    class VistaComiteCreate(BaseModel):
        nombre: str
        descripcion: Optional[str] = None
        es_publica: bool = False
        informes_seleccionados: List[str] = []
        grupos_seleccionados: List[str] = []
        informes_adicionales: List[str] = []
        agrupar_por: str = "informe"
        severidades: List[str] = ["Critica", "Alta", "Media", "Baja"]
    
    class VistaComiteUpdate(BaseModel):
        nombre: Optional[str] = None
        descripcion: Optional[str] = None
        es_publica: Optional[bool] = None
        informes_seleccionados: Optional[List[str]] = None
        grupos_seleccionados: Optional[List[str]] = None
        informes_adicionales: Optional[List[str]] = None
        agrupar_por: Optional[str] = None
        severidades: Optional[List[str]] = None
    
    # ==========================================
    # CRUD VISTAS COMITÉ
    # ==========================================
    
    @router.get("/vistas-comite")
    async def listar_vistas_comite(current_user: CurrentUser = Depends(get_current_user)):
        """Listar todas las vistas de comité accesibles para el usuario."""
        query = {
            "$or": [
                {"es_publica": True},
                {"creado_por": current_user.id}
            ]
        }
        vistas = await db.vistas_comite.find(query, {"_id": 0}).sort("nombre", 1).to_list(100)
        return vistas
    
    @router.get("/vistas-comite/{vista_id}")
    async def obtener_vista_comite(
        vista_id: str,
        current_user: CurrentUser = Depends(get_current_user)
    ):
        """Obtener una vista específica por ID."""
        vista = await db.vistas_comite.find_one({"id": vista_id}, {"_id": 0})
        if not vista:
            raise HTTPException(status_code=404, detail="Vista no encontrada")
        
        # Verificar acceso
        if not vista.get("es_publica") and vista.get("creado_por") != current_user.id:
            raise HTTPException(status_code=403, detail="No tiene acceso a esta vista")
        
        return vista
    
    @router.post("/vistas-comite")
    async def crear_vista_comite(
        data: VistaComiteCreate,
        current_user: CurrentUser = Depends(get_current_user)
    ):
        """Crear una nueva vista de comité."""
        # Verificar nombre único
        existing = await db.vistas_comite.find_one({
            "nombre": {"$regex": f"^{data.nombre}$", "$options": "i"}
        })
        if existing:
            raise HTTPException(status_code=400, detail="Ya existe una vista con ese nombre")
        
        doc = {
            "id": str(uuid.uuid4()),
            "nombre": data.nombre,
            "descripcion": data.descripcion,
            "es_publica": data.es_publica,
            "informes_seleccionados": data.informes_seleccionados,
            "grupos_seleccionados": data.grupos_seleccionados,
            "informes_adicionales": data.informes_adicionales,
            "agrupar_por": data.agrupar_por,
            "severidades": data.severidades,
            "creado_por": current_user.id,
            "creado_por_nombre": current_user.username,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.vistas_comite.insert_one(doc)
        if "_id" in doc:
            del doc["_id"]
        return doc
    
    @router.put("/vistas-comite/{vista_id}")
    async def actualizar_vista_comite(
        vista_id: str,
        data: VistaComiteUpdate,
        current_user: CurrentUser = Depends(get_current_user)
    ):
        """Actualizar una vista de comité existente."""
        vista = await db.vistas_comite.find_one({"id": vista_id}, {"_id": 0})
        if not vista:
            raise HTTPException(status_code=404, detail="Vista no encontrada")
        
        # Solo el creador o admin puede editar
        if vista.get("creado_por") != current_user.id and not current_user.es_admin:
            raise HTTPException(status_code=403, detail="No tiene permisos para editar esta vista")
        
        # Construir update
        update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        if data.nombre is not None:
            # Verificar nombre único (excluyendo la vista actual)
            existing = await db.vistas_comite.find_one({
                "nombre": {"$regex": f"^{data.nombre}$", "$options": "i"},
                "id": {"$ne": vista_id}
            })
            if existing:
                raise HTTPException(status_code=400, detail="Ya existe otra vista con ese nombre")
            update_data["nombre"] = data.nombre
        
        if data.descripcion is not None:
            update_data["descripcion"] = data.descripcion
        if data.es_publica is not None:
            update_data["es_publica"] = data.es_publica
        if data.informes_seleccionados is not None:
            update_data["informes_seleccionados"] = data.informes_seleccionados
        if data.grupos_seleccionados is not None:
            update_data["grupos_seleccionados"] = data.grupos_seleccionados
        if data.informes_adicionales is not None:
            update_data["informes_adicionales"] = data.informes_adicionales
        if data.agrupar_por is not None:
            update_data["agrupar_por"] = data.agrupar_por
        if data.severidades is not None:
            update_data["severidades"] = data.severidades
        
        await db.vistas_comite.update_one({"id": vista_id}, {"$set": update_data})
        
        updated = await db.vistas_comite.find_one({"id": vista_id}, {"_id": 0})
        return updated
    
    @router.delete("/vistas-comite/{vista_id}")
    async def eliminar_vista_comite(
        vista_id: str,
        current_user: CurrentUser = Depends(get_current_user)
    ):
        """Eliminar una vista de comité."""
        vista = await db.vistas_comite.find_one({"id": vista_id}, {"_id": 0})
        if not vista:
            raise HTTPException(status_code=404, detail="Vista no encontrada")
        
        # Solo el creador o admin puede eliminar
        if vista.get("creado_por") != current_user.id and not current_user.es_admin:
            raise HTTPException(status_code=403, detail="No tiene permisos para eliminar esta vista")
        
        await db.vistas_comite.delete_one({"id": vista_id})
        return {"message": "Vista eliminada exitosamente"}
    
    # ==========================================
    # EXPORTACIÓN EXCEL EJECUTIVA
    # ==========================================
    
    @router.get("/vistas-comite/{vista_id}/exportar-excel")
    async def exportar_vista_comite_excel(
        vista_id: str,
        current_user: CurrentUser = Depends(get_current_user)
    ):
        """
        Exportar Vista Comité a Excel con formato ejecutivo.
        
        Columnas:
        1. Nombre de reporte
        2. Fecha Reporte (Mes Año)
        3. Críticas (P/T)
        4. Altas (P/T)
        5. Total Vulns Altas (suma de Críticas + Altas totales)
        6. Estado remediación (desglose con viñetas)
        7. Responsable (lista con viñetas)
        8. Fecha de compromiso
        9. Tiempo de retraso (meses)
        """
        # Obtener vista
        vista = await db.vistas_comite.find_one({"id": vista_id}, {"_id": 0})
        if not vista:
            raise HTTPException(status_code=404, detail="Vista no encontrada")
        
        # Verificar acceso
        if not vista.get("es_publica") and vista.get("creado_por") != current_user.id:
            raise HTTPException(status_code=403, detail="No tiene acceso a esta vista")
        
        # Obtener datos agregados
        datos = await _obtener_datos_comite(
            db=db,
            informes=vista.get("informes_seleccionados", []),
            grupos=vista.get("grupos_seleccionados", []),
            informes_adicionales=vista.get("informes_adicionales", []),
            agrupar_por=vista.get("agrupar_por", "informe")
        )
        
        # Generar Excel
        excel_buffer = _generar_excel_comite(datos, vista.get("nombre", "Vista Comité"))
        
        filename = f"Vista_Comite_{vista.get('nombre', 'Reporte').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @router.get("/vista-comite/exportar-excel")
    async def exportar_vista_comite_excel_directo(
        informes: str = "",
        grupos: str = "",
        informes_adicionales: str = "",
        agrupar_por: str = "informe",
        current_user: CurrentUser = Depends(get_current_user)
    ):
        """
        Exportar Vista Comité a Excel directamente con parámetros (sin vista guardada).
        """
        informes_list = [i.strip() for i in informes.split(",") if i.strip()] if informes else []
        grupos_list = [g.strip() for g in grupos.split(",") if g.strip()] if grupos else []
        informes_adicionales_list = [i.strip() for i in informes_adicionales.split(",") if i.strip()] if informes_adicionales else []
        
        if not informes_list and not grupos_list:
            raise HTTPException(status_code=400, detail="Debe seleccionar al menos un informe o grupo")
        
        # Obtener datos agregados
        datos = await _obtener_datos_comite(
            db=db,
            informes=informes_list,
            grupos=grupos_list,
            informes_adicionales=informes_adicionales_list,
            agrupar_por=agrupar_por
        )
        
        # Generar Excel
        excel_buffer = _generar_excel_comite(datos, "Vista Comité")
        
        filename = f"Vista_Comite_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    # ==========================================
    # FUNCIONES AUXILIARES
    # ==========================================
    
    async def _obtener_datos_comite(db, informes, grupos, informes_adicionales, agrupar_por):
        """Obtener datos agregados para Vista Comité con información extendida."""
        
        # Expandir grupos a informes
        informes_list = list(informes)
        grupo_to_informes = {}
        
        if agrupar_por == "grupo" and grupos:
            grupos_docs = await db.grupos_informes.find(
                {"$or": [{"id": {"$in": grupos}}, {"nombre": {"$in": grupos}}]},
                {"_id": 0}
            ).to_list(1000)
            
            for g in grupos_docs:
                grupo_to_informes[g["nombre"]] = g.get("informes", [])
                informes_list.extend(g.get("informes", []))
        
        # Agregar informes adicionales
        if informes_adicionales:
            informes_list.extend(informes_adicionales)
        
        informes_list = list(set(informes_list))
        
        if not informes_list:
            return []
        
        # Obtener todas las vulnerabilidades
        vulns = await db.vulnerabilidades.find(
            {"nombre_informe_pentest": {"$in": informes_list}},
            {
                "_id": 0, 
                "nombre_informe_pentest": 1, 
                "severidad": 1, 
                "estatus": 1, 
                "responsable": 1, 
                "fecha_hallazgo": 1,
                "fecha_compromiso": 1
            }
        ).to_list(50000)
        
        # Obtener metadatos de informes (fecha del reporte)
        informes_meta = await db.informes_pentest.find(
            {"nombre": {"$in": informes_list}},
            {"_id": 0, "nombre": 1, "fecha": 1}
        ).to_list(500)
        informe_fecha_map = {i["nombre"]: i.get("fecha") for i in informes_meta}
        
        # Mapeo informe -> grupo
        informe_to_grupo = {}
        if agrupar_por == "grupo":
            for grupo_nombre, grupo_informes in grupo_to_informes.items():
                for inf in grupo_informes:
                    informe_to_grupo[inf] = grupo_nombre
        
        # Estados cerrados
        closed_statuses = ["Cerrado", "Corregido", "Desestimado"]
        retest_statuses = ["Para Re Test"]
        
        # Agregar por informe/grupo
        informe_data = defaultdict(lambda: {
            "criticas_pendientes": 0, "criticas_total": 0,
            "altas_pendientes": 0, "altas_total": 0,
            "medias_pendientes": 0, "medias_total": 0,
            "bajas_pendientes": 0, "bajas_total": 0,
            "en_retest": 0,
            "responsables": set(),
            "fecha_mas_antigua": None,
            "fecha_compromiso_max": None,
            "fecha_reporte": None
        })
        
        for v in vulns:
            informe = v.get("nombre_informe_pentest", "Sin informe")
            severidad = v.get("severidad", "")
            estatus = v.get("estatus", "")
            responsable = v.get("responsable")
            fecha_hallazgo = v.get("fecha_hallazgo")
            fecha_compromiso = v.get("fecha_compromiso")
            
            # Determinar key de agrupación
            if agrupar_por == "grupo":
                key = informe_to_grupo.get(informe, informe)
            else:
                key = informe
            
            data = informe_data[key]
            
            # Fecha del reporte
            if not data["fecha_reporte"] and informe in informe_fecha_map:
                data["fecha_reporte"] = informe_fecha_map[informe]
            
            # Responsables
            if responsable:
                data["responsables"].add(responsable)
            
            # Fecha más antigua
            if fecha_hallazgo:
                try:
                    if data["fecha_mas_antigua"] is None or fecha_hallazgo < data["fecha_mas_antigua"]:
                        data["fecha_mas_antigua"] = fecha_hallazgo
                except:
                    pass
            
            # Fecha compromiso máxima
            if fecha_compromiso:
                try:
                    if data["fecha_compromiso_max"] is None or fecha_compromiso > data["fecha_compromiso_max"]:
                        data["fecha_compromiso_max"] = fecha_compromiso
                except:
                    pass
            
            # Contadores
            is_pending = estatus not in closed_statuses
            is_retest = estatus in retest_statuses
            
            if is_retest:
                data["en_retest"] += 1
            
            if severidad == "Critica":
                data["criticas_total"] += 1
                if is_pending:
                    data["criticas_pendientes"] += 1
            elif severidad == "Alta":
                data["altas_total"] += 1
                if is_pending:
                    data["altas_pendientes"] += 1
            elif severidad == "Media":
                data["medias_total"] += 1
                if is_pending:
                    data["medias_pendientes"] += 1
            elif severidad == "Baja":
                data["bajas_total"] += 1
                if is_pending:
                    data["bajas_pendientes"] += 1
        
        # Construir resultado
        result = []
        today = datetime.now(timezone.utc).date()
        
        for key in sorted(informe_data.keys()):
            data = informe_data[key]
            
            # Calcular tiempo de retraso en meses
            tiempo_retraso_meses = 0
            if data["fecha_mas_antigua"]:
                try:
                    fecha_str = data["fecha_mas_antigua"]
                    fecha_date = datetime.strptime(fecha_str[:10], "%Y-%m-%d").date()
                    months_diff = (today.year - fecha_date.year) * 12 + (today.month - fecha_date.month)
                    tiempo_retraso_meses = max(0, months_diff)
                except:
                    pass
            
            # Formato fecha reporte (Mes Año)
            fecha_reporte_fmt = "ND"
            if data["fecha_reporte"]:
                try:
                    fecha_dt = datetime.strptime(data["fecha_reporte"][:10], "%Y-%m-%d")
                    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
                    fecha_reporte_fmt = f"{meses[fecha_dt.month - 1]} {fecha_dt.year}"
                except:
                    pass
            
            # Estado remediación (desglose)
            estado_lines = []
            if data["criticas_pendientes"] > 0:
                estado_lines.append(f"• {data['criticas_pendientes']} Críticas")
            if data["altas_pendientes"] > 0:
                estado_lines.append(f"• {data['altas_pendientes']} Altas pendientes")
            if data["en_retest"] > 0:
                estado_lines.append(f"• {data['en_retest']} En retest")
            if not estado_lines:
                estado_lines.append("• Sin pendientes")
            
            estado_remediacion = "\n".join(estado_lines)
            
            # Responsables con viñetas
            responsables_list = sorted(data["responsables"])
            responsables_fmt = "\n".join([f"• {r}" for r in responsables_list]) if responsables_list else "• ND"
            
            # Fecha compromiso
            fecha_compromiso_fmt = "ND"
            if data["fecha_compromiso_max"]:
                try:
                    fecha_dt = datetime.strptime(data["fecha_compromiso_max"][:10], "%Y-%m-%d")
                    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
                    fecha_compromiso_fmt = f"{meses[fecha_dt.month - 1]} {fecha_dt.year}"
                except:
                    pass
            
            result.append({
                "nombre_reporte": f"•{key}",
                "fecha_reporte": fecha_reporte_fmt,
                "criticas_pendientes": data["criticas_pendientes"],
                "criticas_total": data["criticas_total"],
                "altas_pendientes": data["altas_pendientes"],
                "altas_total": data["altas_total"],
                "total_vulns_altas": data["criticas_total"] + data["altas_total"],
                "estado_remediacion": estado_remediacion,
                "responsable": responsables_fmt,
                "fecha_compromiso": fecha_compromiso_fmt,
                "tiempo_retraso_meses": tiempo_retraso_meses
            })
        
        return result
    
    def _generar_excel_comite(datos, titulo):
        """Generar archivo Excel con formato ejecutivo para Vista Comité."""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Vista Comité"
        
        # ==========================================
        # ESTILOS
        # ==========================================
        
        # Color verde corporativo para cabecera
        verde_corp = PatternFill(start_color="4A7C31", end_color="4A7C31", fill_type="solid")
        
        # Fuentes
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_normal = Font(name="Calibri", size=10)
        font_bold = Font(name="Calibri", size=10, bold=True)
        font_red = Font(name="Calibri", size=10, color="8B0000")  # Rojo oscuro/marrón
        font_red_bold = Font(name="Calibri", size=10, bold=True, color="8B0000")
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='C0C0C0'),
            right=Side(style='thin', color='C0C0C0'),
            top=Side(style='thin', color='C0C0C0'),
            bottom=Side(style='thin', color='C0C0C0')
        )
        
        # Alineaciones
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # ==========================================
        # CABECERAS
        # ==========================================
        
        headers = [
            ("Nombre de reporte", 35),
            ("Fecha Reporte", 12),
            ("Críticas\n(P/T)", 10),
            ("Altas\n(P/T)", 10),
            ("Total\nVulns Altas", 12),
            ("Estado\nremediación", 22),
            ("Responsable", 18),
            ("Fecha de\ncompromiso", 14),
            ("Tiempo de\nretraso\n(meses)", 12)
        ]
        
        # Escribir cabeceras
        for col_idx, (header_text, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = font_header
            cell.fill = verde_corp
            cell.alignment = align_center
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Altura de cabecera
        ws.row_dimensions[1].height = 45
        
        # ==========================================
        # DATOS
        # ==========================================
        
        for row_idx, item in enumerate(datos, 2):
            # Col 1: Nombre de reporte (izquierda)
            cell = ws.cell(row=row_idx, column=1, value=item["nombre_reporte"])
            cell.font = font_normal
            cell.alignment = align_left
            cell.border = thin_border
            
            # Col 2: Fecha Reporte (centro)
            cell = ws.cell(row=row_idx, column=2, value=item["fecha_reporte"])
            cell.font = font_normal
            cell.alignment = align_center
            cell.border = thin_border
            
            # Col 3: Críticas P/T (centro, rojo si hay pendientes)
            criticas_text = f"{item['criticas_pendientes']}/{item['criticas_total']}"
            cell = ws.cell(row=row_idx, column=3, value=criticas_text)
            if item["criticas_pendientes"] > 0:
                cell.font = font_red_bold
            else:
                cell.font = font_normal
            cell.alignment = align_center
            cell.border = thin_border
            
            # Col 4: Altas P/T (centro, rojo si hay pendientes)
            altas_text = f"{item['altas_pendientes']}/{item['altas_total']}"
            cell = ws.cell(row=row_idx, column=4, value=altas_text)
            if item["altas_pendientes"] > 0:
                cell.font = font_red_bold
            else:
                cell.font = font_normal
            cell.alignment = align_center
            cell.border = thin_border
            
            # Col 5: Total Vulns Altas (centro, negrita)
            cell = ws.cell(row=row_idx, column=5, value=item["total_vulns_altas"])
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = thin_border
            
            # Col 6: Estado remediación (izquierda, multilínea)
            cell = ws.cell(row=row_idx, column=6, value=item["estado_remediacion"])
            cell.font = font_normal
            cell.alignment = align_left
            cell.border = thin_border
            
            # Col 7: Responsable (izquierda, multilínea)
            cell = ws.cell(row=row_idx, column=7, value=item["responsable"])
            cell.font = font_normal
            cell.alignment = align_left
            cell.border = thin_border
            
            # Col 8: Fecha compromiso (centro)
            cell = ws.cell(row=row_idx, column=8, value=item["fecha_compromiso"])
            cell.font = font_normal
            cell.alignment = align_center
            cell.border = thin_border
            
            # Col 9: Tiempo retraso (centro)
            cell = ws.cell(row=row_idx, column=9, value=item["tiempo_retraso_meses"])
            cell.font = font_normal
            cell.alignment = align_center
            cell.border = thin_border
            
            # Ajustar altura de fila según contenido
            num_lines = max(
                item["estado_remediacion"].count("\n") + 1,
                item["responsable"].count("\n") + 1,
                1
            )
            ws.row_dimensions[row_idx].height = max(25, num_lines * 15)
        
        # ==========================================
        # GUARDAR
        # ==========================================
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    return router

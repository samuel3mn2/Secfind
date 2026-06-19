"""
Test suite for Vista Comité CRUD + Excel Export (Iteration 22)
- POST /api/vistas-comite
- GET /api/vistas-comite, GET /api/vistas-comite/{id}
- PUT /api/vistas-comite/{id} (update without duplicate)
- DELETE /api/vistas-comite/{id}
- GET /api/vistas-comite/{id}/exportar-excel
"""
import io
import os
import uuid
import pytest
import requests
from openpyxl import load_workbook


def _read_env_url():
    url = os.environ.get('REACT_APP_BACKEND_URL', '').strip()
    if url:
        return url.rstrip('/')
    try:
        with open('/app/frontend/.env') as f:
            for line in f:
                if line.startswith('REACT_APP_BACKEND_URL='):
                    return line.split('=', 1)[1].strip().rstrip('/')
    except Exception:
        pass
    return ''


BASE_URL = _read_env_url()
assert BASE_URL, "REACT_APP_BACKEND_URL must be set"


@pytest.fixture(scope="module")
def auth_headers():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": "admin", "password": "admin123"})
    assert r.status_code == 200, f"Login failed: {r.text}"
    return {"Authorization": f"Bearer {r.json()['token']}"}


@pytest.fixture(scope="module")
def informes_list(auth_headers):
    r = requests.get(f"{BASE_URL}/api/dropdown-options", headers=auth_headers)
    if r.status_code != 200:
        return []
    return r.json().get("informes_pentest", [])


@pytest.fixture
def created_vista(auth_headers, informes_list):
    nombre = f"TEST_VC_{uuid.uuid4().hex[:8]}"
    payload = {
        "nombre": nombre,
        "descripcion": "Vista creada en test",
        "es_publica": False,
        "informes_seleccionados": informes_list[:2] if informes_list else [],
        "grupos_seleccionados": [],
        "informes_adicionales": [],
        "agrupar_por": "informe",
        "severidades": ["Critica", "Alta"]
    }
    r = requests.post(f"{BASE_URL}/api/vistas-comite", json=payload, headers=auth_headers)
    assert r.status_code == 200, f"Create failed: {r.text}"
    data = r.json()
    yield data
    # Cleanup
    requests.delete(f"{BASE_URL}/api/vistas-comite/{data['id']}", headers=auth_headers)


class TestVistasComiteCRUD:
    def test_auth_required(self):
        r = requests.get(f"{BASE_URL}/api/vistas-comite")
        assert r.status_code in (401, 403)

    def test_list_vistas(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/vistas-comite", headers=auth_headers)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_create_vista(self, created_vista):
        assert "id" in created_vista
        assert created_vista["nombre"].startswith("TEST_VC_")
        assert created_vista["agrupar_por"] == "informe"
        assert "_id" not in created_vista
        assert isinstance(created_vista["informes_seleccionados"], list)

    def test_create_duplicate_name_rejected(self, auth_headers, created_vista):
        payload = {
            "nombre": created_vista["nombre"],
            "informes_seleccionados": []
        }
        r = requests.post(f"{BASE_URL}/api/vistas-comite", json=payload, headers=auth_headers)
        assert r.status_code == 400

    def test_get_vista_by_id(self, auth_headers, created_vista):
        r = requests.get(f"{BASE_URL}/api/vistas-comite/{created_vista['id']}", headers=auth_headers)
        assert r.status_code == 200
        data = r.json()
        assert data["id"] == created_vista["id"]
        assert data["nombre"] == created_vista["nombre"]
        assert "_id" not in data

    def test_get_vista_not_found(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/vistas-comite/nonexistent-id", headers=auth_headers)
        assert r.status_code == 404

    def test_update_vista_no_duplicate(self, auth_headers, created_vista, informes_list):
        """PUT should update existing doc (no new doc) and persist changes."""
        # Count vistas before
        before = requests.get(f"{BASE_URL}/api/vistas-comite", headers=auth_headers).json()
        before_count = len(before)

        new_nombre = created_vista["nombre"] + "_UPD"
        new_desc = "Descripcion actualizada"
        new_informes = informes_list[:1] if informes_list else []
        update_payload = {
            "nombre": new_nombre,
            "descripcion": new_desc,
            "informes_seleccionados": new_informes
        }
        r = requests.put(
            f"{BASE_URL}/api/vistas-comite/{created_vista['id']}",
            json=update_payload,
            headers=auth_headers
        )
        assert r.status_code == 200, r.text
        updated = r.json()
        assert updated["id"] == created_vista["id"]
        assert updated["nombre"] == new_nombre
        assert updated["descripcion"] == new_desc
        assert updated["informes_seleccionados"] == new_informes

        # Verify persistence via GET
        r2 = requests.get(f"{BASE_URL}/api/vistas-comite/{created_vista['id']}", headers=auth_headers)
        assert r2.status_code == 200
        fetched = r2.json()
        assert fetched["nombre"] == new_nombre
        assert fetched["descripcion"] == new_desc

        # Verify NO duplicate doc created (same count)
        after = requests.get(f"{BASE_URL}/api/vistas-comite", headers=auth_headers).json()
        assert len(after) == before_count, "PUT must not create a duplicate vista"

    def test_update_duplicate_name_rejected(self, auth_headers, created_vista):
        # Create second vista
        other_nombre = f"TEST_VC_{uuid.uuid4().hex[:8]}"
        r2 = requests.post(
            f"{BASE_URL}/api/vistas-comite",
            json={"nombre": other_nombre, "informes_seleccionados": []},
            headers=auth_headers
        )
        assert r2.status_code == 200
        other_id = r2.json()["id"]

        try:
            # Try to rename created_vista to other_nombre -> should 400
            r = requests.put(
                f"{BASE_URL}/api/vistas-comite/{created_vista['id']}",
                json={"nombre": other_nombre},
                headers=auth_headers
            )
            assert r.status_code == 400
        finally:
            requests.delete(f"{BASE_URL}/api/vistas-comite/{other_id}", headers=auth_headers)

    def test_delete_vista(self, auth_headers, informes_list):
        # Create a fresh vista to delete
        nombre = f"TEST_VC_DEL_{uuid.uuid4().hex[:8]}"
        r = requests.post(
            f"{BASE_URL}/api/vistas-comite",
            json={"nombre": nombre, "informes_seleccionados": []},
            headers=auth_headers
        )
        assert r.status_code == 200
        vista_id = r.json()["id"]

        rd = requests.delete(f"{BASE_URL}/api/vistas-comite/{vista_id}", headers=auth_headers)
        assert rd.status_code == 200

        # Verify 404
        rg = requests.get(f"{BASE_URL}/api/vistas-comite/{vista_id}", headers=auth_headers)
        assert rg.status_code == 404


class TestVistaComiteExcelExport:
    def test_export_excel_returns_valid_xlsx(self, auth_headers, created_vista):
        url = f"{BASE_URL}/api/vistas-comite/{created_vista['id']}/exportar-excel"
        r = requests.get(url, headers=auth_headers)
        assert r.status_code == 200, r.text
        ctype = r.headers.get("content-type", "")
        assert "spreadsheetml" in ctype or "officedocument" in ctype, f"Unexpected content-type: {ctype}"
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".xlsx" in cd
        # Validate it's a real xlsx
        wb = load_workbook(io.BytesIO(r.content))
        assert "Vista Comité" in wb.sheetnames or wb.active.title == "Vista Comité"

    def test_export_excel_has_9_columns_and_headers(self, auth_headers, created_vista):
        url = f"{BASE_URL}/api/vistas-comite/{created_vista['id']}/exportar-excel"
        r = requests.get(url, headers=auth_headers)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active

        expected = [
            "Nombre de reporte",
            "Fecha Reporte",
            "Críticas",
            "Altas",
            "Total",
            "Estado",
            "Responsable",
            "Fecha de",
            "Tiempo de",
        ]
        for col in range(1, 10):
            cell_value = ws.cell(row=1, column=col).value or ""
            assert expected[col - 1].lower() in cell_value.lower(), \
                f"Column {col} header '{cell_value}' missing expected token '{expected[col-1]}'"

        # Ensure exactly 9 header columns populated
        assert ws.cell(row=1, column=10).value in (None, "")

    def test_export_excel_header_styles(self, auth_headers, created_vista):
        url = f"{BASE_URL}/api/vistas-comite/{created_vista['id']}/exportar-excel"
        r = requests.get(url, headers=auth_headers)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active

        cell = ws.cell(row=1, column=1)
        # Verde corporativo 4A7C31
        fill_color = (cell.fill.start_color.rgb or "").upper()
        assert "4A7C31" in fill_color, f"Header fill color {fill_color} does not contain 4A7C31"
        # Font white bold
        font_color = (cell.font.color.rgb or "").upper() if cell.font.color else ""
        assert "FFFFFF" in font_color, f"Header font color is {font_color}"
        assert cell.font.bold is True
        # Border (thin)
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"

    def test_export_excel_404_for_invalid_id(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/vistas-comite/nonexistent-id-xyz/exportar-excel",
            headers=auth_headers
        )
        assert r.status_code == 404


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
